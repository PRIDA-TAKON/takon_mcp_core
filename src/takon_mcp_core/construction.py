import os
import glob
import time
import base64
import requests
import io
import json
from datetime import datetime
from typing import Optional, List
from mcp.server.fastmcp import FastMCP
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from dotenv import load_dotenv
from PIL import Image

load_dotenv()

# Initialize MCP for Construction
mcp = FastMCP("Takon Construction")

# Global variable to cache local model
LOCAL_VISION_MODEL = None

def get_local_vision_model():
    """Lazy load moondream model for local vision."""
    global LOCAL_VISION_MODEL
    if LOCAL_VISION_MODEL is None:
        try:
            import moondream as md
            LOCAL_VISION_MODEL = md.vl(model="moondream-latest")
            print("Local Vision Model (Moondream) loaded successfully.")
        except Exception as e:
            print(f"Error loading local vision model: {e}")
    return LOCAL_VISION_MODEL

# --- AI Helpers ---
def call_local_ai(prompt: str, image_path: Optional[str] = None) -> str:
    """Run AI locally using Moondream (Vision) or a text fallback."""
    model = get_local_vision_model()
    if model is None: return "Error: Local model not available."
    try:
        if image_path:
            image = Image.open(image_path)
            return model.query(image, prompt)["answer"].strip()
        else:
            # For text-only, Moondream 2 can do basic captioning/QA, 
            # but for general summary we might need a small LLM or use Moondream's internal cap.
            return model.caption(Image.new('RGB', (100, 100)))["caption"] # Mock or simple cap
    except Exception as e:
        return f"Local AI Error: {str(e)}"

def call_ai_api(system_prompt: str, user_prompt: str, image_base64: Optional[str] = None, image_path: Optional[str] = None) -> str:
    """Universal AI Call: Supports Local (Moondream) and API (Typhoon, OpenAI, Ollama)."""
    strategy = os.getenv("AI_STRATEGY", "api")
    
    if strategy == "local":
        return call_local_ai(f"{system_prompt}. {user_prompt}", image_path)

    api_key = os.getenv("AI_API_KEY") or os.getenv("TYPHOON_API_KEY")
    base_url = os.getenv("AI_BASE_URL", "https://api.opentyphoon.ai/v1")
    model = os.getenv("AI_VISION_MODEL", "typhoon-ocr") if image_base64 else os.getenv("AI_TEXT_MODEL", "typhoon-v2.1-12b-instruct")

    if not api_key and "localhost" not in base_url: return "Error: API Key missing."

    headers = {"Authorization": f"Bearer {api_key}" if api_key else "", "Content-Type": "application/json"}
    content = [{"type": "text", "text": user_prompt}]
    if image_base64: content.append({"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{image_base64}"}})
    
    payload = {
        "model": model,
        "messages": [{"role": "system", "content": system_prompt}, {"role": "user", "content": content}],
        "max_tokens": 1024, "temperature": 0.4
    }
    try:
        response = requests.post(f"{base_url.rstrip('/')}/chat/completions", headers=headers, json=payload, timeout=45)
        if response.status_code == 200: return response.json()["choices"][0]["message"]["content"].strip()
    except Exception as e: return f"Error: {str(e)}"
    return "Error: AI Call failed."

# --- Tools ---
@mcp.tool()
def generate_daily_report(project_dir: str, report_date: str) -> str:
    """Generate a daily report with AI analysis (Local or API)."""
    template_path = os.path.join(project_dir, "ข้อมูลพื้นฐาน", "REPORT_template_00.xlsx")
    if not os.path.exists(template_path):
        template_path = "REPORT_template_00.xlsx"
        if not os.path.exists(template_path): return "Error: Template not found."

    picture_dir = os.path.join(project_dir, "รูปถ่ายจากพื้นที่ก่อสร้าง", "รูปถ่ายยังไม่แยก")
    output_dir = os.path.join(project_dir, "รายงานประจำวัน")
    if not os.path.exists(output_dir): os.makedirs(output_dir)

    try:
        thai_date = datetime.strptime(report_date, "%Y-%m-%d").strftime("%d/%m/") + str(datetime.strptime(report_date, "%Y-%m-%d").year + 543)
    except: return "Error: Date format YYYY-MM-DD."

    output_file = os.path.join(output_dir, f"DailyReport_{report_date}.xlsx")
    image_slots = [{'cell': 'E66', 'text': 'E67'}, {'cell': 'Z66', 'text': 'Z67'}, {'cell': 'E69', 'text': 'E70'}, {'cell': 'Z69', 'text': 'Z70'}, {'cell': 'E72', 'text': 'E73'}, {'cell': 'Z72', 'text': 'Z73'}]
    
    try:
        wb = load_workbook(template_path); ws = wb.active 
        ws['AK7'] = thai_date; ws['Y10'] = "08:00 - 17:00"
        images = glob.glob(os.path.join(picture_dir, "*.jpg")) + glob.glob(os.path.join(picture_dir, "*.png"))
        daily_captions = []

        for idx in range(min(6, len(images))):
            img_path = images[idx]; img = OpenpyxlImage(img_path); slot = image_slots[idx]
            img.width, img.height = 500, 375; ws.add_image(img, slot['cell'])
            with open(img_path, "rb") as f: b64 = base64.b64encode(f.read()).decode('utf-8')
            caption = call_ai_api("คุณคือวิศวกรประเมินหน้างาน ให้เลือก 1 หมวดหมู่: งานโครงผนัง, งานตกแต่งภายใน, งานฝ้าเพดาน, งานระบบปรับอากาศ, งานระบบไฟฟ้า, งานทำความสะอาด, งานสี, งานฉาบผนัง, งานสุขภัณฑ์, งานเตรียมพื้นที่, งานแก้ไขงาน", "ตอบเพียงชื่อหมวดหมู่", image_base64=b64, image_path=img_path)
            ws[slot['text']] = caption; daily_captions.append(caption)

        ws['N34'] = call_ai_api("คุณคือวิศวกรคุมงาน สรุปงานสั้นๆ", f"สรุปจาก: {', '.join(daily_captions)}", image_path=None) if daily_captions else "ไม่มีงานวันนี้"
        wb.save(output_file); return f"Success: Report saved at {output_file}"
    except Exception as e: return f"Error: {str(e)}"

@mcp.tool()
def generate_project_dashboard(project_dir: str) -> str:
    """Generate an HTML Dashboard with AI insights from all reports (Local or API)."""
    boq_path = os.path.join(project_dir, "ข้อมูลพื้นฐาน", "BOQ.xlsx")
    reports_dir = os.path.join(project_dir, "รายงานประจำวัน")
    output_path = os.path.join(project_dir, "แผนงานและความคืบหน้า", "dashboard.html")
    if not os.path.exists(os.path.dirname(output_path)): os.makedirs(os.path.dirname(output_path))

    # 1. Load BOQ
    items = []
    if os.path.exists(boq_path):
        try:
            wb = load_workbook(boq_path, data_only=True); ws = wb.active
            for r in range(2, ws.max_row + 1):
                if ws[f'C{r}'].value: items.append({"desc": ws[f'C{r}'].value, "prog": 0})
        except: pass

    # 2. Get AI Insights from reports
    reports = glob.glob(os.path.join(reports_dir, "*.xlsx"))
    all_summaries = []
    for r in reports[:10]: # Look at last 10 reports
        try:
            wb = load_workbook(r, data_only=True); all_summaries.append(wb.active['N34'].value or "")
        except: pass
    
    ai_insight = call_ai_api("คุณคือที่ปรึกษาโครงการ สรุปภาพรวมความคืบหน้าและคำแนะนำจากข้อมูลรายงานเหล่านี้", f"ข้อมูลรายงาน: {' | '.join(all_summaries)}", image_path=None)

    # 3. Build HTML
    html = f"""
    <!DOCTYPE html><html><head><meta charset="utf-8"><title>Project Dashboard</title>
    <style>body{{font-family:sans-serif; background:#f0f2f5; padding:20px;}} .card{{background:white; padding:20px; border-radius:10px; box-shadow:0 2px 5px rgba(0,0,0,0.1); margin-bottom:20px;}}
    .bar-bg{{background:#eee; height:20px; border-radius:10px;}} .bar{{background:#4caf50; height:100%; border-radius:10px;}} h1{{color:#1a73e8;}}</style></head>
    <body><h1>🏗️ Project Dashboard</h1>
    <div class="card"><h2>🤖 AI Insight (Analysis)</h2><p>{ai_insight}</p></div>
    <div class="card"><h2>📋 Work Items</h2>
    {''.join([f"<div><p>{i['desc']} ({i['prog']}%)</p><div class='bar-bg'><div class='bar' style='width:{i['prog']}%'></div></div></div>" for i in items])}
    </div></body></html>
    """
    with open(output_path, "w", encoding="utf-8") as f: f.write(html)
    return f"Success: Dashboard generated at {output_path}"

def main(): mcp.run()
if __name__ == "__main__": main()
